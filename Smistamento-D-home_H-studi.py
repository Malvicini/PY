import os
import shutil
import re
from openpyxl import load_workbook

# Percorsi
excel_file = "elenco_codici_studi.xlsx"
cartella_sorgente = r"D:\home\CheckOut"
cartella_destinazione = r"H:\96-GESTIONE_STUDI\DOWNLOAD"

# Carica file Excel
wb = load_workbook(excel_file)
ws = wb.active

def normalizza(testo):
    """Rende il testo confrontabile (minuscolo e senza caratteri speciali)"""
    return re.sub(r'[^a-z0-9]', '', testo.lower())

# Lista file presenti nella cartella sorgente
file_presenti = os.listdir(cartella_sorgente)

print("=== AVVIO ELABORAZIONE ===\n")

# Loop sulle righe Excel
for row in ws.iter_rows(min_row=2):  # salta intestazione
    codice_cella = row[0]  # colonna A
    stato_cella = row[1]   # colonna B

    codice = codice_cella.value

    if not codice:
        continue

    codice_norm = normalizza(str(codice))
    trovato = False

    print(f"\nControllo codice: {codice}")

    for file in file_presenti:
        nome_file = normalizza(file)

        if codice_norm in nome_file:
            origine = os.path.join(cartella_sorgente, file)
            destinazione = os.path.join(cartella_destinazione, file)

            try:
                shutil.move(origine, destinazione)
                print(f"  ✔ Spostato: {file}")
                print(f"    Da: {origine}")
                print(f"    A:  {destinazione}")
                trovato = True
            except Exception as e:
                print(f"  ❌ Errore nello spostamento di {file}: {e}")

    if trovato:
        stato_cella.value = "scaricato"
        print("  → Stato aggiornato: scaricato")
    else:
        print("  → Nessun file trovato")

# Salva Excel
wb.save(excel_file)

print("\n=== OPERAZIONE COMPLETATA ===")