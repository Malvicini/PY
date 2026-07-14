import getpass
import os
import re
from datetime import date

import openpyxl
import pandas as pd

from config import DEFAULT_DRAWINGS_DIR
from helpers import sanitize_filesystem_name
from pdf_finder import _derive_family_prefix


class DataLoader:
    """Load families and sequences from the provided Excel file.

    Expectations:
    - Sheet1 contains families with columns: family_code, family_name
    - Sheet2 contains sequences with columns: sequence_id, family_code, description
    """

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        # Lazy load caches
        self._families = None
        self._sequences = None
        self._groups_machines = None

    def _load(self):
        # Read sheets
        xls = pd.ExcelFile(self.excel_path)
        # Heuristics: first sheet -> families, second -> sequences
        try:
            df_fam = pd.read_excel(xls, sheet_name=0)
        except Exception:
            df_fam = pd.DataFrame()
        try:
            df_seq = pd.read_excel(xls, sheet_name=1)
        except Exception:
            df_seq = pd.DataFrame()

        # Normalize column names to lowercase
        df_fam.columns = [c.lower() for c in df_fam.columns]
        df_seq.columns = [c.lower() for c in df_seq.columns]

        self._families = df_fam
        self._sequences = df_seq

    def get_families(self):
        if self._families is None:
            self._load()
        df = self._families
        if df is None or df.empty:
            return []
        # Try to find appropriate columns (prefer columns containing 'cod' for short code)
        code_col = None
        id_col = None
        name_col = None
        for c in df.columns:
            lc = c.lower()
            if 'cod' in lc and code_col is None:
                code_col = c
            if 'id' == lc or lc.endswith('id') or lc == 'id':
                id_col = c
            if 'name' in lc or 'descr' in lc or 'des' in lc:
                name_col = c

        # Fallbacks
        if id_col is None and len(df.columns) > 0:
            id_col = df.columns[0]
        if code_col is None:
            # prefer second column if first is id
            code_col = df.columns[1] if len(df.columns) > 1 else id_col
        if name_col is None:
            # prefer description-like column else code
            name_col = df.columns[2] if len(df.columns) > 2 else code_col

        families = []
        for _, row in df.iterrows():
            families.append({
                'family_id': str(row.get(id_col, '')).strip(),
                'family_code': str(row.get(code_col, '')).strip(),
                'family_name': str(row.get(name_col, '')).strip(),
            })
        return families

    def get_sequences_for_family(self, family_code: str):
        if self._sequences is None:
            self._load()
        df = self._sequences
        if df is None or df.empty:
            return []
        # For your workbook, sheet 'Studi' uses column 'CODICE' for family short code
        # and 'PROGRESSIVO' for the sequence number. Prefer these explicitly.
        cols = [c.lower() for c in df.columns]
        fam_col = None
        prog_col = None
        desc_col = None

        for c in df.columns:
            lc = c.lower()
            if lc == 'codice' or 'cod' in lc:
                fam_col = c
            if lc == 'progressivo' or 'prog' in lc:
                prog_col = c
            if 'descr' in lc or 'description' in lc or 'desc' in lc:
                desc_col = c

        # Fallbacks
        if fam_col is None:
            fam_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if prog_col is None:
            # use PROGRESSIVO if present else use ID or first col
            prog_col = next((c for c in df.columns if 'progress' in c.lower() or 'prog' in c.lower()), df.columns[0])
        if desc_col is None:
            desc_col = next((c for c in df.columns if 'descr' in c.lower() or 'description' in c.lower() or 'des' in c.lower()), prog_col)

        fc = str(family_code).strip()
        filtered = df[df[fam_col].astype(str).str.strip() == fc]

        sequences = []
        for _, row in filtered.iterrows():
            seq_id = str(row.get(prog_col, '')).strip()
            # Pad sequence_id to 3 digits with leading zeros
            try:
                seq_id = str(int(seq_id)).zfill(3)
            except ValueError:
                pass  # keep as is if not numeric
            sequences.append({
                'sequence_id': seq_id,
                'description': str(row.get(desc_col, '')).strip(),
            })
        return sequences

    def get_next_progressivo(self, family_code: str):
        if self._sequences is None:
            self._load()
        df = self._sequences
        if df is None or df.empty:
            return '001'

        fam_col = None
        prog_col = None
        for c in df.columns:
            lc = c.lower()
            if lc == 'codice' or 'cod' in lc and fam_col is None:
                fam_col = c
            if lc == 'progressivo' or 'prog' in lc:
                prog_col = c
        if fam_col is None:
            fam_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if prog_col is None:
            prog_col = next((c for c in df.columns if 'progress' in c.lower() or 'prog' in c.lower()), df.columns[2] if len(df.columns) > 2 else df.columns[0])

        filtered = df[df[fam_col].astype(str).str.strip().str.upper() == str(family_code).strip().upper()]
        max_prog = 0
        for _, row in filtered.iterrows():
            value = row.get(prog_col, '')
            if pd.isna(value) or str(value).strip() == '':
                continue
            try:
                item = int(str(value).strip())
                max_prog = max(max_prog, item)
            except ValueError:
                continue
        return str(max_prog + 1).zfill(3)

    def _ensure_study_drawings_folder(self, family_code: str, full_code: str, base_dir: str = None):
        drawings_base = base_dir or os.environ.get('DRAWINGS_DIR', DEFAULT_DRAWINGS_DIR)
        if not drawings_base:
            return None

        drawings_base = os.path.abspath(os.path.expanduser(str(drawings_base)))
        os.makedirs(drawings_base, exist_ok=True)

        safe_code = sanitize_filesystem_name(str(full_code).strip())
        if not safe_code:
            return None

        safe_family = sanitize_filesystem_name(str(family_code).strip())
        if not safe_family:
            safe_family = _derive_family_prefix(safe_code)
        if not safe_family:
            safe_family = safe_code

        # The folder name should follow the family code from the Excel workbook.
        # Example: GS-U008 belongs under the family folder GS-U, not under GS.
        if safe_family and safe_code.upper().startswith(f'{safe_family.upper()}'):
            safe_family = safe_family

        family_dir = os.path.join(drawings_base, safe_family)
        os.makedirs(family_dir, exist_ok=True)

        study_dir = os.path.join(family_dir, safe_code)
        os.makedirs(study_dir, exist_ok=True)

        return study_dir

    def create_new_study(self, family_code: str, description: str, user: str = None, created_date: str = None):
        if created_date is None:
            created_date = date.today()
        if user is None:
            try:
                user = os.environ.get('USERNAME') or os.environ.get('USER') or getpass.getuser() or ''
            except Exception:
                user = ''

        family_code = str(family_code).strip()
        if not family_code:
            raise ValueError('family_code is required')

        valid_codes = {str(f.get('family_code', '')).strip().upper() for f in self.get_families()}
        if family_code.upper() not in valid_codes:
            raise ValueError(f'family_code "{family_code}" is not a valid family')

        wb = openpyxl.load_workbook(self.excel_path)
        if 'Studi' not in wb.sheetnames:
            raise ValueError('Sheet "Studi" not found in workbook')
        ws = wb['Studi']

        headers = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in ws[1]]
        header_index = {name: idx + 1 for idx, name in enumerate(headers)}

        idx_id = header_index.get('id', 1)
        idx_codice = header_index.get('codice', 2)
        idx_progressivo = header_index.get('progressivo', 3)
        idx_descrizione = header_index.get('descrizione', 4)
        idx_data = header_index.get('data', 5)
        idx_utente = header_index.get('utente', 6)
        idx_datamodifica = header_index.get('datamodifica', None)
        idx_utentemodifica = header_index.get('utentemodifica', None)
        idx_colonna4 = header_index.get('colonna4', 12)

        progressivo = self.get_next_progressivo(family_code)
        full_code = f'{family_code}{progressivo}'

        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=idx_id, value='-')
        ws.cell(row=target_row, column=idx_codice, value=family_code)
        ws.cell(row=target_row, column=idx_progressivo, value=progressivo)
        ws.cell(row=target_row, column=idx_descrizione, value=description or '')
        ws.cell(row=target_row, column=idx_data, value=created_date)
        ws.cell(row=target_row, column=idx_utente, value=user or '')
        if idx_datamodifica:
            ws.cell(row=target_row, column=idx_datamodifica, value=created_date)
        if idx_utentemodifica:
            ws.cell(row=target_row, column=idx_utentemodifica, value=user or '')
        ws.cell(row=target_row, column=idx_colonna4, value=full_code)

        wb.save(self.excel_path)

        try:
            drawings_folder = self._ensure_study_drawings_folder(family_code, full_code)
        except Exception:
            drawings_folder = None

        self._sequences = None
        return {
            'family_code': family_code,
            'progressivo': progressivo,
            'description': description or '',
            'date': created_date.strftime('%d/%m/%Y') if hasattr(created_date, 'strftime') else str(created_date),
            'user': user or '',
            'full_code': full_code,
            'drawings_folder': drawings_folder,
        }

    def get_groups_machines(self):
        if self._groups_machines is None:
            self._load_groups_machines()
        return self._groups_machines

    def _load_groups_machines(self):
        # Read the third sheet for groups and machines details
        xls = pd.ExcelFile(self.excel_path)
        try:
            df = pd.read_excel(xls, sheet_name=2)  # Third sheet
        except Exception:
            df = pd.DataFrame()

        # Normalize column names to lowercase
        df.columns = [c.lower() for c in df.columns]

        self._groups_machines = df

    def get_groups_machines_for_family(self, family_code: str, sequence_code: str = None):
        if self._groups_machines is None:
            self._load_groups_machines()
        df = self._groups_machines
        if df is None or df.empty:
            return []

        # Filter by family code
        fc = str(family_code).strip().upper()
        filtered = df[df['cod'].astype(str).str.strip().str.upper() == fc]

        if sequence_code:
            # Filter by sequence code (progressivo), normalize by removing leading zeros
            sc_normalized = str(sequence_code).lstrip('0') or '0'
            filtered = filtered[filtered['pro'].astype(str).str.strip().str.lstrip('0').fillna('0') == sc_normalized]

        groups_machines = []
        for _, row in filtered.iterrows():
            groups_machines.append({
                'id': str(row.get('id', '')).strip(),
                'cod': str(row.get('cod', '')).strip(),
                'pro': str(row.get('pro', '')).strip(),
                'tipo': str(row.get('tipo', '')).strip(),
                'articolo': str(row.get('articolo', '')).strip(),
                'desart': str(row.get('desart', '')).strip(),
            })
        return groups_machines
