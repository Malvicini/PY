#!/usr/bin/env python3
"""Run macro_1.py for each study code and mark downloads.

Behavior:
- Reads study codes from `DATA_FILE` sheet `SHEET_NAME` column A (configurable).
- For each code (starting after the last saved index), runs `macro_1.py` and sends the code as stdin (so VAR1 is filled).
- After running the macro, waits for a new file to appear in `DOWNLOAD_DIR`. When a new file appears, writes "scaricato" into column B for that row.
- Press the `End` key to stop the loop gracefully. Progress is saved to `STATE_FILE` so a subsequent run resumes after the last marked row.

Configure constants below to match your workbook/sheet layout.
"""
import os
import sys
import time
import json
import threading
import subprocess
from pathlib import Path

import openpyxl
from pynput import keyboard

# Configuration - edit as needed
DATA_FILE = Path('elenco_codici_studi.xlsx')
SHEET_NAME = None  # None -> use active sheet or first
CODE_COLUMN = 'A'  # column containing the study code (one per row)
MARK_COLUMN = 'B'  # column to write 'scaricato'
START_ROW = 2      # first row with data (assumes header in row 1)
MACRO_SCRIPT = Path('macro_1.py')
DOWNLOAD_DIR = Path(r'H:\96-GESTIONE_STUDI\DOWNLOAD')
STATE_FILE = Path('.runner_state.json')
PYTHON = sys.executable  # run macro with same python interpreter
# Stop key configuration: can be 'end', 'ctrl'/'control', or a single character like 'c'
STOP_KEY = 'ctrl'

# Runtime flags
stop_flag = False


def on_press(key):
    global stop_flag
    try:
        # If STOP_KEY is 'end', check special Key.end
        if isinstance(STOP_KEY, str) and STOP_KEY.lower() in ('end', 'endkey'):
            if key == keyboard.Key.end:
                print('\nStop key (End) pressed — stopping after current iteration...')
                stop_flag = True
                return False

        # If STOP_KEY is ctrl/control, check control keys
        if isinstance(STOP_KEY, str) and STOP_KEY.lower() in ('ctrl', 'control'):
            if key in (keyboard.Key.ctrl_l, keyboard.Key.ctrl_r):
                print('\nStop key (Ctrl) pressed — stopping after current iteration...')
                stop_flag = True
                return False

        # If STOP_KEY is a single character, check alphanumeric key presses
        if isinstance(STOP_KEY, str) and len(STOP_KEY) == 1:
            try:
                if getattr(key, 'char', None) and key.char.lower() == STOP_KEY.lower():
                    print(f"\nStop key ('{STOP_KEY}') pressed — stopping after current iteration...")
                    stop_flag = True
                    return False
            except Exception:
                pass
    except Exception:
        pass


def load_state():
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def save_state(d):
    STATE_FILE.write_text(json.dumps(d), encoding='utf-8')


def read_codes(workbook_path, sheet_name=None):
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    codes = []
    rows = []
    for r in range(START_ROW, ws.max_row + 1):
        cell = ws[f"{CODE_COLUMN}{r}"]
        val = cell.value
        if val is None:
            continue
        code = str(val).strip()
        if code == '':
            continue
        codes.append(code)
        rows.append(r)
    return wb, ws, codes, rows


def mark_download(ws, row, mark_text='scaricato'):
    ws[f"{MARK_COLUMN}{row}"] = mark_text


def check_for_new_file(code, before_set):
    """Check if a new file appeared matching the code.
    
    Returns the new file path if found, None otherwise.
    Does NOT block — just a single check.
    """
    now = set(os.listdir(DOWNLOAD_DIR)) if DOWNLOAD_DIR.exists() else set()
    new = now - before_set
    
    if not new:
        print(f'No new files for {code}')
        return None
    
    # Find the first new file that was not there before
    for fname in sorted(new):
        fpath = DOWNLOAD_DIR / fname
        # Check if the code appears anywhere in the filename
        if code.lower() in fname.lower():
            print(f'Found new file matching code: {fname}')
            return fpath
    
    # If no exact match found, just take the first new file
    # (fallback: assume the first new file is the download)
    fname = sorted(new)[0]
    print(f'Found new file (no code match, but first new): {fname}')
    return DOWNLOAD_DIR / fname


def run_macro_for_code(code):
    # Run macro and pass code to stdin so macro's input() receives it
    print(f'Running macro for: {code}')
    proc = subprocess.run([PYTHON, str(MACRO_SCRIPT)], input=code + '\n', text=True)
    return proc.returncode


def main():
    global stop_flag
    if not DATA_FILE.exists():
        print('Data file not found:', DATA_FILE)
        return

    wb, ws, codes, rows = read_codes(DATA_FILE, SHEET_NAME)
    if not codes:
        print('No codes found in', DATA_FILE)
        return

    state = load_state()
    last_index = state.get('last_index', -1)
    start_idx = last_index + 1
    print(f'Loaded {len(codes)} codes. Starting at index {start_idx}.')

    # Start keyboard listener to listen for End key
    kb = keyboard.Listener(on_press=on_press)
    kb.start()

    for idx in range(start_idx, len(codes)):
        if stop_flag:
            print('Stop requested; exiting loop.')
            break

        code = codes[idx]
        row = rows[idx]

        before = set(os.listdir(DOWNLOAD_DIR)) if DOWNLOAD_DIR.exists() else set()

        # run macro (will prompt for VAR1 but we feed code via stdin)
        try:
            rc = run_macro_for_code(code)
            if rc != 0:
                print(f'Macro exited with code {rc} for {code}. Skipping file check and continuing...')
                # Skip the file check if macro failed
                state['last_index'] = idx
                save_state(state)
                time.sleep(1)
                continue
        except Exception as e:
            print(f'Macro failed with exception for {code}: {e}. Continuing...')
            state['last_index'] = idx
            save_state(state)
            time.sleep(1)
            continue

        # Check for new file (non-blocking, single check)
        newfile = check_for_new_file(code, before)
        if newfile and newfile.exists():
            print('Detected new file:', newfile.name)
            mark_download(ws, row, 'scaricato')
            wb.save(DATA_FILE)
        else:
            print(f'No new file found for {code}')
        
        # update state regardless
        state['last_index'] = idx
        save_state(state)

        # small pause between iterations
        time.sleep(1)

    kb.stop()
    print('Runner finished.')


if __name__ == '__main__':
    main()
